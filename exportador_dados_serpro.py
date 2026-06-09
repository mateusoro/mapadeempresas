import sys
import io

# UTF-8 output para Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from datetime import datetime
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# Mudar para o diretório do script (suporta execução via caminho completo)
SCRIPT_DIR = Path(__file__).parent.resolve()
os.chdir(SCRIPT_DIR)

# Configuração de pasta para logs de snapshots
logs_folder = os.path.join(os.getcwd(), "serpro_snapshots_logs") + os.sep
if not os.path.exists(logs_folder):
    os.makedirs(logs_folder)
    print(f"Pasta de logs criada: {logs_folder}")

# Configuração do Chrome
options = webdriver.ChromeOptions()
prefs = {
    "download.default_directory": os.path.join(os.getcwd(), "downloads_serpro"),
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
options.add_experimental_option("prefs", prefs)

# Criar pasta de downloads se não existir
download_folder = os.path.join(os.getcwd(), "downloads_serpro")
if not os.path.exists(download_folder):
    os.makedirs(download_folder)
    print(f"Pasta de downloads criada: {download_folder}")

# Configuração do banco de dados SQLite
cnxn = sqlite3.connect('base_dados.db')

def salvar_snapshot(nome_etapa):
    """Salva snapshot HTML da página para auditoria"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    try:
        # Salvar HTML completo
        html_content = driver.page_source
        html_filename = os.path.join(logs_folder, f"{timestamp}_{nome_etapa}.html")
        with open(html_filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"  [HTML] {html_filename}")
        
        # Salvar log resumido
        log_filename = os.path.join(logs_folder, f"{timestamp}_{nome_etapa}_log.txt")
        with open(log_filename, 'w', encoding='utf-8') as f:
            f.write(f"Etapa: {nome_etapa}\n")
            f.write(f"Timestamp: {timestamp}\n")
            f.write(f"URL: {driver.current_url}\n")
            f.write(f"Titulo: {driver.title}\n")
        print(f"  [LOG] {log_filename}")
        
        return True
    except Exception as e:
        print(f"✗ Erro ao salvar snapshot: {e}")
        return False

def aguardar_carregamento_pagina(timeout=60):
    """Aguarda a página carregar completamente"""
    print("\n[1] Aguardando carregamento da página...")
    
    try:
        # Espera o document estar ready
        WebDriverWait(driver, timeout).until(
            lambda driver: driver.execute_script("return document.readyState") == "complete"
        )
        print("✓ Página carregada (document.readyState = complete)")
        
        # Aguarda um pouco extra para elementos renderizarem
        time.sleep(3)
        
        salvar_snapshot("01_pagina_carregada")
        return True
        
    except Exception as e:
        print(f"✗ Erro no carregamento: {e}")
        salvar_snapshot("01_erro_carregamento")
        return False

def aguardar_elemento_qv_inner_object(intervalo=10, timeout_max=600):
    """
    Aguarda o elemento <div class="qv-inner-object"> aparecer
    Verifica a cada 'intervalo' segundos até 'timeout_max'
    """
    print("\n[1.5] Aguardando elemento 'qv-inner-object'...")
    print(f"       Verificando a cada {intervalo} segundos (timeout: {timeout_max}s)")
    
    tempo_decorrido = 0
    tentativa = 0
    
    while tempo_decorrido < timeout_max:
        tentativa += 1
        
        try:
            # Verifica se o elemento existe
            elemento = driver.find_element(By.CLASS_NAME, "qv-inner-object")
            
            if elemento:
                print(f"✓ Elemento encontrado na tentativa {tentativa} ({tempo_decorrido}s)")
                salvar_snapshot("01_5_qv_inner_object_encontrado")
                return True
                
        except:
            pass
        
        # Verifica também via CSS selector (mais específico)
        try:
            elemento = driver.find_element(By.CSS_SELECTOR, "div.qv-inner-object")
            if elemento:
                print(f"✓ Elemento encontrado via CSS selector na tentativa {tentativa} ({tempo_decorrido}s)")
                salvar_snapshot("01_5_qv_inner_object_encontrado")
                return True
        except:
            pass
        
        # Se não encontrou, aguarda e verifica novamente
        print(f"  [{tentativa}] Não encontrado. Aguardando {intervalo}s... ({tempo_decorrido}/{timeout_max}s)")
        time.sleep(intervalo)
        tempo_decorrido += intervalo
    
    print(f"✗ Timeout! Elemento não encontrado em {timeout_max} segundos")
    salvar_snapshot("01_5_erro_qv_inner_object_timeout")
    return False

def clicar_botao_direito_meio_tela():
    """Clica com botão direito no meio da tela e aguarda o menu aparecer"""
    print("\n[2] Clicando com botão direito no meio da tela...")
    
    try:
        # Obtém dimensões da janela
        window_size = driver.get_window_size()
        meio_x = window_size['width'] // 2
        meio_y = window_size['height'] // 2
        
        # Tenta encontrar um elemento SVG ou canvas (comum em dashboards)
        elementos_interativos = driver.find_elements(By.CSS_SELECTOR, "svg, canvas, [role='main'], .object-wrapper, [ng-app]")
        
        if elementos_interativos:
            elemento = elementos_interativos[0]
            print(f"✓ Elemento interativo encontrado: {elemento.tag_name}")
        else:
            elemento = driver.find_element(By.TAG_NAME, "body")
            print("  Usando elemento body")
        
        # Clica com botão direito
        actions = ActionChains(driver)
        actions.context_click(elemento).perform()
        
        print(f"✓ Clique direito realizado em ({meio_x}, {meio_y})")
        
        # Aguarda o menu aparecer
        time.sleep(2)
        
        salvar_snapshot("02_menu_contexto")
        return True
        
    except Exception as e:
        print(f"✗ Erro ao clicar com botão direito: {e}")
        salvar_snapshot("02_erro_menu_contexto")
        return False

def clicar_exportar_dados():
    """Clica na opção 'Exportar dados' via JavaScript"""
    print("\n[3] Clicando em 'Exportar dados'...")
    
    try:
        resultado = driver.execute_script("""
            const elementos = Array.from(document.querySelectorAll('*'));
            const encontrado = elementos.find(el => 
                el.textContent.toLowerCase().includes('exportar') && 
                el.textContent.toLowerCase().includes('dados')
            );
            if (encontrado) {
                encontrado.click();
                return true;
            }
            return false;
        """)
        
        if resultado:
            print("✓ Clicado em 'Exportar dados'")
            time.sleep(3)  # Aguarda menu aparecer
            salvar_snapshot("03_exportar_dados_clicado")
            return True
        
        print("✗ 'Exportar dados' não encontrado")
        salvar_snapshot("03_erro_exportar_dados")
        return False
        
    except Exception as e:
        print(f"✗ Erro ao clicar em 'Exportar dados': {e}")
        salvar_snapshot("03_erro_exportar_dados")
        return False
        
    except Exception as e:
        print(f"✗ Erro ao clicar em 'Exportar dados': {e}")
        salvar_snapshot("03_erro_exportar_dados")
        return False

def clicar_formatacao_tabela():
    """Clica em 'Formatação de tabela' via JavaScript"""
    print("\n[4] Clicando em 'Formatação de tabela'...")
    
    try:
        resultado = driver.execute_script("""
            const elementos = Array.from(document.querySelectorAll('*'));
            const encontrado = elementos.find(el => 
                el.textContent.toLowerCase().includes('formatação') && 
                el.textContent.toLowerCase().includes('tabela')
            ) || elementos.find(el => 
                el.textContent.toLowerCase().includes('formato')
            );
            if (encontrado) {
                encontrado.click();
                return true;
            }
            return false;
        """)
        
        if resultado:
            print("✓ Clicado em 'Formatação de tabela'")
            time.sleep(2)
            salvar_snapshot("04_formatacao_tabela_clicada")
            return True
        
        print("✗ 'Formatação de tabela' não encontrado")
        salvar_snapshot("04_erro_formatacao_tabela")
        return False
        
    except Exception as e:
        print(f"✗ Erro ao clicar em 'Formatação de tabela': {e}")
        salvar_snapshot("04_erro_formatacao_tabela")
        return False
        
    except Exception as e:
        print(f"✗ Erro: {e}")
        salvar_snapshot("04_erro_formatacao_tabela")
        return False

def aguardar_botao_export_url(intervalo=5, timeout_max=120):
    """
    Aguarda o botão <a class="export-url"> aparecer e clica nele
    Este é o link para download que aparece APÓS clicar em "Exportar"
    """
    print("\n[5.5] Aguardando link de download (botão export-url)...")
    print(f"       Verificando a cada {intervalo} segundos (timeout: {timeout_max}s)")
    
    tempo_decorrido = 0
    tentativa = 0
    
    while tempo_decorrido < timeout_max:
        tentativa += 1
        
        try:
            # Procura pelo elemento <a class="export-url">
            elemento = driver.find_element(By.CSS_SELECTOR, "a.export-url")
            
            if elemento and elemento.is_displayed():
                print(f"✓ Botão 'export-url' encontrado na tentativa {tentativa} ({tempo_decorrido}s)")
                print(f"  URL: {elemento.get_attribute('href')[:80]}...")
                
                # Scroll para o elemento
                driver.execute_script("arguments[0].scrollIntoView(true);", elemento)
                time.sleep(0.5)
                
                # Clica no botão
                elemento.click()
                print(f"✓ Clique no botão export-url realizado!")
                
                salvar_snapshot("05_5_botao_export_clicado")
                time.sleep(2)  # Aguarda o navegador processar o clique
                
                return True
        
        except Exception as e:
            pass
        
        # Se não encontrou, aguarda e verifica novamente
        print(f"  [{tentativa}] Link não encontrado. Aguardando... ({tempo_decorrido}/{timeout_max}s)")
        time.sleep(intervalo)
        tempo_decorrido += intervalo
    
    print(f"✗ Timeout! Botão export-url não apareceu em {timeout_max} segundos")
    salvar_snapshot("05_5_erro_export_url_timeout")
    return False

def verificar_download_completo(intervalo=10, timeout_max=600):
    """
    Verifica se um arquivo foi baixado na pasta downloads_serpro
    Aguarda a cada 'intervalo' segundos até 'timeout_max'
    """
    print("\n[5.6] Verificando download do arquivo...")
    print(f"       Verificando a cada {intervalo} segundos (timeout: {timeout_max}s)")
    print(f"       Pasta: {download_folder}")
    
    tempo_decorrido = 0
    tentativa = 0
    arquivo_encontrado = None
    
    while tempo_decorrido < timeout_max:
        tentativa += 1
        
        try:
            # Lista arquivos na pasta de downloads
            arquivos = os.listdir(download_folder)
            
            if arquivos:
                # Filtra arquivos que não são .crdownload (downloads incompletos)
                arquivos_completos = [f for f in arquivos if not f.endswith('.crdownload') and not f.endswith('.tmp')]
                
                if arquivos_completos:
                    arquivo_encontrado = arquivos_completos[0]
                    tamanho = os.path.getsize(os.path.join(download_folder, arquivo_encontrado))
                    print(f"✓ Arquivo baixado na tentativa {tentativa} ({tempo_decorrido}s)")
                    print(f"  Nome: {arquivo_encontrado}")
                    print(f"  Tamanho: {tamanho} bytes")
                    salvar_snapshot("05_5_download_completo")
                    return True
                else:
                    print(f"  [{tentativa}] Arquivo ainda baixando... ({tempo_decorrido}/{timeout_max}s)")
            else:
                print(f"  [{tentativa}] Nenhum arquivo detectado. Aguardando... ({tempo_decorrido}/{timeout_max}s)")
        
        except Exception as e:
            print(f"  Erro ao verificar pasta: {e}")
        
        time.sleep(intervalo)
        tempo_decorrido += intervalo
    
    print(f"✗ Timeout! Arquivo não foi baixado em {timeout_max} segundos")
    salvar_snapshot("05_5_erro_download_timeout")
    return False

def clicar_exportar_final():
    """Clica no botão 'Exportar' final via JavaScript"""
    print("\n[5] Clicando em 'Exportar'...")
    
    try:
        resultado = driver.execute_script("""
            const botoes = Array.from(document.querySelectorAll('button, [role="button"]'));
            const encontrado = botoes.find(el => 
                el.textContent.toLowerCase().includes('exportar')
            );
            if (encontrado) {
                encontrado.click();
                return true;
            }
            return false;
        """)
        
        if resultado:
            print("✓ Clicado em 'Exportar'")
            salvar_snapshot("05_exportacao_clicada")
            time.sleep(2)
            
            # Aguarda o link de download aparecer
            if not aguardar_botao_export_url(intervalo=5, timeout_max=120):
                return False
            
            # Aguarda o download completar
            if verificar_download_completo(intervalo=10, timeout_max=600):
                salvar_snapshot("05_exportacao_completada")
                return True
            else:
                return False
        
        print("✗ Botão 'Exportar' não encontrado")
        salvar_snapshot("05_erro_exportacao")
        return False
        
    except Exception as e:
        print(f"✗ Erro ao clicar em 'Exportar': {e}")
        salvar_snapshot("05_erro_exportacao")
        return False
        
        # [ETAPA FINAL] Aguarda o download completar
        if verificar_download_completo(intervalo=10, timeout_max=600):
            salvar_snapshot("05_exportacao_completada")
            return True
        else:
            return False
        
    except Exception as e:
        print(f"✗ Erro ao clicar: {e}")
        print("  Tentando alternativa via JavaScript...")
        
        try:
            resultado = driver.execute_script("""
                const botoes = Array.from(document.querySelectorAll('button, [role="button"]'));
                const encontrado = botoes.find(el => 
                    el.textContent.toLowerCase().includes('exportar')
                );
                if (encontrado) {
                    encontrado.click();
                    return true;
                }
                return false;
            """)
            
            if resultado:
                print("✓ Clicado via JavaScript")
                salvar_snapshot("05_exportacao_clicada")
                time.sleep(1)
                
                # Aguarda o link e o download
                if not aguardar_botao_export_url(intervalo=5, timeout_max=120):
                    return False
                
                if verificar_download_completo(intervalo=10, timeout_max=600):
                    salvar_snapshot("05_exportacao_completada")
                    return True
                else:
                    return False
        
        except Exception as e2:
            print(f"✗ Alternativa também falhou: {e2}")
        
        salvar_snapshot("05_erro_exportacao")
        return False

def criar_tabela_brutos_se_nao_existe(cnxn, nome_tabela_brutos='dados_brutos_mes'):
    """Cria tabela temporaria para dados BRUTOS do mes atual (com des_secao).

    Esta tabela guarda o .xlsx importado de um mes especifico ANTES da
    agregacao. Sempre e limpa (DELETE) antes de cada importacao mensal.
    """
    cur = cnxn.cursor()
    cur.execute(f'''
        CREATE TABLE IF NOT EXISTS {nome_tabela_brutos} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ano_abertura TEXT,
            mes_abertura TEXT,
            ano_baixa TEXT,
            mes_baixa TEXT,
            regiao TEXT,
            uf TEXT,
            municipio TEXT,
            natureza_juridica TEXT,
            des_secao TEXT,
            tipo_situacao TEXT,
            porte TEXT,
            opcao_mei TEXT,
            quantidade INTEGER
        )
    ''')
    cnxn.commit()
    print(f"  [TEMP] Tabela '{nome_tabela_brutos}' pronta")


def limpar_tabela_brutos(cnxn, nome_tabela_brutos='dados_brutos_mes'):
    """Limpa a tabela temporaria de dados brutos. Chamada antes de cada import."""
    cur = cnxn.cursor()
    cur.execute(f"DELETE FROM {nome_tabela_brutos}")
    cnxn.commit()
    print(f"  [TEMP] Tabela '{nome_tabela_brutos}' limpa")


def agregar_e_inserir(cnxn,
                       nome_tabela_brutos='dados_brutos_mes',
                       nome_tabela_destino='dados_serpro',
                       ano_abertura=None, mes_abertura=None):
    """Agrega dados brutos do mes SOMANDO quantidade por 12 colunas (sem des_secao).

    Para cada (ano, mes) lido da tabela temporaria, executa:
        INSERT OR IGNORE INTO dados_serpro (12 cols + quantidade)
        SELECT 12 cols, SUM(quantidade)
        FROM dados_brutos_mes
        WHERE ano_abertura=? AND mes_abertura=?
        GROUP BY 12 cols

    Se a UNIQUE constraint de dados_serpro ja tem 12 colunas (versao nova),
    a soma agrega naturalmente. Se ainda tem 13 colunas (versao antiga),
    a soma de varias linhas identicas nas 13 colunas (incluindo des_secao)
    ainda funciona, mas nao economiza espaco.
    """
    cur = cnxn.cursor()
    if ano_abertura is None or mes_abertura is None:
        return False
    # Descobre as colunas da tabela destino via PRAGMA
    cur.execute(f"PRAGMA table_info({nome_tabela_destino})")
    cols_destino = [r[1] for r in cur.fetchall() if r[1] != 'id']

    # Se a destino tem des_secao, precisamos agregar 12 colunas.
    if 'des_secao' in cols_destino:
        # Modo legado: destino ainda tem des_secao, entao inserimos cada linha
        # da temp com INSERT OR IGNORE. Sem agregacao real.
        cur.execute(f'''
            INSERT OR IGNORE INTO {nome_tabela_destino}
                (ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                 regiao, uf, municipio, natureza_juridica,
                 des_secao, tipo_situacao, porte, opcao_mei, quantidade)
            SELECT
                ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                regiao, uf, municipio, natureza_juridica,
                des_secao, tipo_situacao, porte, opcao_mei, quantidade
            FROM {nome_tabela_brutos}
            WHERE ano_abertura = ? AND mes_abertura = ?
        ''', (str(ano_abertura), str(mes_abertura)))
        inseridos = cur.rowcount
        cnxn.commit()
        return inseridos > 0

    # Modo novo: destino tem 12 colunas (sem des_secao) + quantidade
    cur.execute(f'''
        INSERT OR IGNORE INTO {nome_tabela_destino}
            (ano_abertura, mes_abertura, ano_baixa, mes_baixa,
             regiao, uf, municipio, natureza_juridica,
             tipo_situacao, porte, opcao_mei, quantidade)
        SELECT
            ano_abertura, mes_abertura, ano_baixa, mes_baixa,
            regiao, uf, municipio, natureza_juridica,
            tipo_situacao, porte, opcao_mei,
            SUM(quantidade) AS quantidade
        FROM {nome_tabela_brutos}
        WHERE ano_abertura = ? AND mes_abertura = ?
        GROUP BY
            ano_abertura, mes_abertura, ano_baixa, mes_baixa,
            regiao, uf, municipio, natureza_juridica,
            tipo_situacao, porte, opcao_mei
    ''', (str(ano_abertura), str(mes_abertura)))
    cnxn.commit()
    inseridos = cur.rowcount
    print(f"  [AGREG] {ano_abertura}/{mes_abertura}: {inseridos:,} grupos unicos inseridos em {nome_tabela_destino}")
    return inseridos


def criar_tabela_dados_serpro_agregada(cnxn, nome_tabela='dados_serpro'):
    """Cria a tabela dados_serpro com schema OTIMIZADO (12 colunas + quantidade).

    A coluna des_secao foi removida. A unique constraint e sobre 12 colunas
    (sem des_secao). A coluna quantidade passa a ser a SOMA das quantidades
    das linhas que compartilham os 12 campos (ou seja, a coluna que era
    'quantidade' no .xlsx vira a soma das des_secao repetidas).
    """
    cur = cnxn.cursor()
    cur.execute(f'''
        CREATE TABLE IF NOT EXISTS {nome_tabela} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ano_abertura TEXT,
            mes_abertura TEXT,
            ano_baixa TEXT,
            mes_baixa TEXT,
            regiao TEXT,
            uf TEXT,
            municipio TEXT,
            natureza_juridica TEXT,
            tipo_situacao TEXT,
            porte TEXT,
            opcao_mei TEXT,
            quantidade INTEGER,
            UNIQUE(
                ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                regiao, uf, municipio, natureza_juridica,
                tipo_situacao, porte, opcao_mei, quantidade
            )
        )
    ''')
    cnxn.commit()
    print(f"  [TABELA] '{nome_tabela}' (12 colunas + id + quantidade) pronta")


def importar_excel_para_sqlite(caminho_arquivo, limpar_antes=False, anos_para_reimportar=None,
                              usar_agregacao=True):
    """Importa o arquivo Excel baixado para o banco SQLite.

    Logica:
        - 1a vez: limpar_antes=True apaga tudo.
        - Runs seguintes:
            * anos_para_reimportar=None (default): so acumula (INSERT OR IGNORE).
            * anos_para_reimportar=[2026]: apaga SOMENTE os registros com
              ano_abertura IN (anos_para_reimportar) antes de inserir. Util
              para o ano atual (que ainda esta crescendo) e qualquer ano
              que voce queira refresh.
            * Anos passados (nao listados) permanecem intactos.

    Quando usar_agregacao=True (default), o import ocorre em 2 fases:
        Fase A: le xlsx -> INSERT INTO dados_brutos_mes (sempre limpa antes)
        Fase B: SELECT com GROUP BY 12 colunas + SUM(quantidade)
                -> INSERT OR IGNORE INTO dados_serpro

    Args:
        caminho_arquivo: caminho do .xlsx
        limpar_antes: se True, faz DELETE FROM antes (modo single-shot).
        anos_para_reimportar: lista de anos a recarregar (ex: [2026]).
        usar_agregacao: se True, passa pela tabela temporaria + agregacao.
                        se False, insere direto (modo legado).
    """
    print("\n" + "="*60)
    print("IMPORTANDO EXCEL PARA SQLITE")
    print("="*60)
    print(f"Arquivo: {caminho_arquivo}")
    if limpar_antes:
        print("Modo: LIMPAR TUDO antes (primeira carga)")
    elif anos_para_reimportar:
        print(f"Modo: REIMPORTAR ANOS {anos_para_reimportar} (demais preservados)")
    else:
        print("Modo: ACUMULAR (preserva tudo, so adiciona)")
    if usar_agregacao:
        print("Agregacao: ATIVA (2 fases: xlsx -> dados_brutos_mes -> dados_serpro)")
    else:
        print("Agregacao: DESATIVADA (modo legado, insere direto)")
    sys.stdout.flush()

    nome_tabela = 'dados_serpro'
    nome_tabela_brutos = 'dados_brutos_mes'

    try:
        # Conectar ao banco
        cnxn = sqlite3.connect('base_dados.db')
        cursor = cnxn.cursor()

        # Criar tabela temporaria de dados brutos (sempre com 13 cols + des_secao)
        criar_tabela_brutos_se_nao_existe(cnxn, nome_tabela_brutos)
        limpar_tabela_brutos(cnxn, nome_tabela_brutos)
        cnxn.commit()

        # Criar tabela se nao existir
        if usar_agregacao:
            # Schema otimizado: 12 cols (sem des_secao) + quantidade
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {nome_tabela} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ano_abertura TEXT,
                    mes_abertura TEXT,
                    ano_baixa TEXT,
                    mes_baixa TEXT,
                    regiao TEXT,
                    uf TEXT,
                    municipio TEXT,
                    natureza_juridica TEXT,
                    tipo_situacao TEXT,
                    porte TEXT,
                    opcao_mei TEXT,
                    quantidade INTEGER,
                    UNIQUE(
                        ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                        regiao, uf, municipio, natureza_juridica,
                        tipo_situacao, porte, opcao_mei, quantidade
                    )
                )
            ''')
            cnxn.commit()
        else:
            # Modo legado: 13 cols (com des_secao)
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {nome_tabela} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ano_abertura TEXT,
                    mes_abertura TEXT,
                    ano_baixa TEXT,
                    mes_baixa TEXT,
                    regiao TEXT,
                    uf TEXT,
                    municipio TEXT,
                    natureza_juridica TEXT,
                    des_secao TEXT,
                    tipo_situacao TEXT,
                    porte TEXT,
                    opcao_mei TEXT,
                    quantidade INTEGER,
                    UNIQUE(
                        ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                        regiao, uf, municipio, natureza_juridica,
                        des_secao, tipo_situacao, porte, opcao_mei, quantidade
                    )
                )
            ''')
            cnxn.commit()

        # Migracao automatica: detecta tabela existente sem UNIQUE correto
        cursor.execute(f"PRAGMA table_info({nome_tabela})")
        cols_atual = [r[1] for r in cursor.fetchall()]

        if usar_agregacao and 'des_secao' in cols_atual:
            print("⚠ Tabela destino ainda tem des_secao (versao antiga).")
            print("  Para ativar agregacao, recrie o banco com o backup.")
        elif not usar_agregacao and 'des_secao' not in cols_atual:
            print("⚠ Tabela destino sem des_secao (versao nova/agregada).")
            print("  Modo legado nao funciona com schema otimizado.")

        cursor.execute(f"PRAGMA index_list({nome_tabela})")
        indices = [r[1] for r in cursor.fetchall()]
        tem_unique = False
        for idx in indices:
            try:
                cursor.execute(f"PRAGMA index_info({idx})")
                cols_idx = [r[2] for r in cursor.fetchall()]
                # A UNIQUE esperada: 12 cols (agregada) ou 13 (legado)
                esperado = 12 if usar_agregacao else 13
                if len(cols_idx) >= esperado:
                    tem_unique = True
                    break
            except Exception:
                pass

        if not tem_unique:
            print("⚠ Tabela sem UNIQUE adequado. Re-criando do zero...")
            cursor.execute(f"ALTER TABLE {nome_tabela} RENAME TO {nome_tabela}_old")
            if usar_agregacao:
                cursor.execute(f'''
                    CREATE TABLE {nome_tabela} (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        ano_abertura TEXT, mes_abertura TEXT,
                        ano_baixa TEXT, mes_baixa TEXT,
                        regiao TEXT, uf TEXT, municipio TEXT,
                        natureza_juridica TEXT,
                        tipo_situacao TEXT, porte TEXT, opcao_mei TEXT,
                        quantidade INTEGER,
                        UNIQUE(
                            ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                            regiao, uf, municipio, natureza_juridica,
                            tipo_situacao, porte, opcao_mei, quantidade
                        )
                    )
                ''')
            else:
                cursor.execute(f'''
                    CREATE TABLE {nome_tabela} (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        ano_abertura TEXT, mes_abertura TEXT,
                        ano_baixa TEXT, mes_baixa TEXT,
                        regiao TEXT, uf TEXT, municipio TEXT,
                        natureza_juridica TEXT, des_secao TEXT,
                        tipo_situacao TEXT, porte TEXT, opcao_mei TEXT,
                        quantidade INTEGER,
                        UNIQUE(
                            ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                            regiao, uf, municipio, natureza_juridica,
                            des_secao, tipo_situacao, porte, opcao_mei, quantidade
                        )
                    )
                ''')
            cnxn.commit()
            print("  ✓ Tabela re-criada (dados antigos perdidos, recomece do backup se necessario)")

        # Politica de limpeza
        if limpar_antes:
            print("Limpando dados da tabela dados_serpro...")
            cursor.execute(f"DELETE FROM {nome_tabela}")
            cnxn.commit()
            print("Dados limpos com sucesso.")
        elif anos_para_reimportar:
            placeholders = ",".join("?" * len(anos_para_reimportar))
            cursor.execute(
                f"DELETE FROM {nome_tabela} WHERE ano_abertura IN ({placeholders})",
                [str(a) for a in anos_para_reimportar]
            )
            removidos = cursor.rowcount
            cnxn.commit()
            print(f"Removidos {removidos} registros dos anos {anos_para_reimportar} para reimportacao.")
        sys.stdout.flush()

        # Processar com openpyxl em chunks
        print("Lendo Excel com openpyxl...")
        sys.stdout.flush()

        wb = load_workbook(filename=caminho_arquivo, read_only=True, data_only=True)
        ws = wb.active

        chunk_size = 5000
        chunk_num = 0
        total_inserido = 0
        total_duplicatas = 0
        row_num = 0
        dados = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1

            try:
                dados.append((
                    str(row[0]) if row[0] is not None else '',
                    str(row[1]) if row[1] is not None else '',
                    str(row[2]) if row[2] is not None else '',
                    str(row[3]) if row[3] is not None else '',
                    str(row[4]) if row[4] is not None else '',
                    str(row[5]) if row[5] is not None else '',
                    str(row[6]) if row[6] is not None else '',
                    str(row[7]) if row[7] is not None else '',
                    str(row[8]) if row[8] is not None else '',
                    str(row[9]) if row[9] is not None else '',
                    str(row[10]) if row[10] is not None else '',
                    str(row[11]) if row[11] is not None else '',
                    int(row[12]) if row[12] is not None else 0
                ))
            except Exception as e:
                pass

            if len(dados) >= chunk_size:
                chunk_num += 1
                print(f"  Inserting chunk {chunk_num}: {len(dados)} rows... (row {row_num})")
                sys.stdout.flush()

                tabela_destino_insert = nome_tabela_brutos if usar_agregacao else nome_tabela
                cursor.executemany(f'''
                    INSERT OR IGNORE INTO {tabela_destino_insert} (
                        ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                        regiao, uf, municipio, natureza_juridica,
                        des_secao, tipo_situacao, porte, opcao_mei, quantidade
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', dados)
                cnxn.commit()
                inseridos_no_chunk = cursor.rowcount if cursor.rowcount >= 0 else len(dados)
                total_inserido += inseridos_no_chunk
                total_duplicatas += (len(dados) - inseridos_no_chunk)
                print(f"  Chunk {chunk_num} -> {tabela_destino_insert}: +{inseridos_no_chunk} novos (duplicatas: {len(dados) - inseridos_no_chunk})")
                sys.stdout.flush()
                dados = []

        # Insert remaining
        if dados:
            chunk_num += 1
            print(f"  Inserting final chunk {chunk_num}: {len(dados)} rows...")
            sys.stdout.flush()

            tabela_destino_insert = nome_tabela_brutos if usar_agregacao else nome_tabela
            cursor.executemany(f'''
                INSERT OR IGNORE INTO {tabela_destino_insert} (
                    ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                    regiao, uf, municipio, natureza_juridica,
                    des_secao, tipo_situacao, porte, opcao_mei, quantidade
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', dados)
            cnxn.commit()
            inseridos_no_chunk = cursor.rowcount if cursor.rowcount >= 0 else len(dados)
            total_inserido += inseridos_no_chunk
            total_duplicatas += (len(dados) - inseridos_no_chunk)
            print(f"  Final chunk: +{inseridos_no_chunk} novos (duplicatas: {len(dados) - inseridos_no_chunk})")
            sys.stdout.flush()

        wb.close()

        # FASE B (agregacao): se usar_agregacao, agrupa 12 colunas + SUM(quantidade)
        if usar_agregacao:
            print("\n[FASE B] Agregando dados brutos por 12 colunas + SUM(quantidade)...")
            sys.stdout.flush()
            # Descobre ano/mes presentes na temp
            cursor.execute(f"SELECT DISTINCT ano_abertura, mes_abertura FROM {nome_tabela_brutos}")
            pares = cursor.fetchall()
            print(f"  [FASE B] {len(pares)} pares (ano, mes) distintos na temp")
            total_agregado = 0
            for (a, m) in pares:
                # Limpa dados do mesmo (ano, mes) ja existentes no destino (se modo reimport)
                if anos_para_reimportar and str(a) in [str(x) for x in anos_para_reimportar]:
                    cursor.execute(
                        f"DELETE FROM {nome_tabela} WHERE ano_abertura=? AND mes_abertura=?",
                        (str(a), str(m))
                    )
                    print(f"  [FASE B] {a}/{m}: removidos do destino para reimport")
                elif limpar_antes and a is not None and m is not None:
                    pass  # ja feito pelo DELETE global acima
                # Roda a agregacao
                cursor.execute(f'''
                    INSERT OR IGNORE INTO {nome_tabela}
                        (ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                         regiao, uf, municipio, natureza_juridica,
                         tipo_situacao, porte, opcao_mei, quantidade)
                    SELECT
                        ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                        regiao, uf, municipio, natureza_juridica,
                        tipo_situacao, porte, opcao_mei,
                        SUM(quantidade) AS quantidade
                    FROM {nome_tabela_brutos}
                    WHERE ano_abertura = ? AND mes_abertura = ?
                    GROUP BY
                        ano_abertura, mes_abertura, ano_baixa, mes_baixa,
                        regiao, uf, municipio, natureza_juridica,
                        tipo_situacao, porte, opcao_mei
                ''', (str(a), str(m)))
                cnxn.commit()
                total_agregado += cursor.rowcount
            print(f"  [FASE B] Total agregado inserido em {nome_tabela}: {total_agregado:,}")
            sys.stdout.flush()

        # Verify total
        cursor.execute(f"SELECT COUNT(*) FROM {nome_tabela}")
        total_final = cursor.fetchone()[0]

        cursor.close()
        cnxn.close()

        print("\n" + "="*60)
        print(f"IMPORTACAO CONCLUIDA!")
        print(f"Linhas lidas do Excel:   {row_num}")
        print(f"Novos registros:         +{total_inserido}")
        print(f"Duplicatas ignoradas:     {total_duplicatas}")
        print(f"Total na tabela:         {total_final}")
        print("="*60)
        sys.stdout.flush()
        return True

    except Exception as e:
        print(f"Erro na importacao: {e}")
        import traceback
        traceback.print_exc()
        return False

def executar_ciclo_completo():
    """Executa ciclo completo: baixar + importar"""
    print("\n" + "="*60)
    print("INICIANDO CICLO COMPLETO (BAIXAR + IMPORTAR)")
    print("="*60)
    
    # Etapa 1: Baixar
    if not baixar_somente():
        print("\n✗ Falha na etapa de download")
        return False
    
    # Etapa 2: Importar
    if not importar_ultimo_download():
        print("\n✗ Falha na etapa de importacao")
        return False
    
    print("\n" + "="*60)
    print("✓ CICLO COMPLETO FINALIZADO COM SUCESSO!")
    print("="*60)
    return True

def executar_em_loop(intervalo_minutos=60, limite_tentativas=None):
    """Executa a automação em loop a cada X minutos"""
    print(f"\n⏰ Modo LOOP ativado: execução a cada {intervalo_minutos} minuto(s)")
    if limite_tentativas:
        print(f"   Limite de tentativas: {limite_tentativas}")
    else:
        print(f"   Limite de tentativas: infinito")
    
    tentativa = 0
    while True:
        tentativa += 1
        
        if limite_tentativas and tentativa > limite_tentativas:
            print(f"\n✓ Limite de tentativas ({limite_tentativas}) atingido. Encerrando.")
            break
        
        print(f"\n{'='*60}")
        print(f"TENTATIVA {tentativa} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}")
        
        resultado = executar_ciclo_completo()
        
        if resultado:
            print(f"\n✓ Sucesso na tentativa {tentativa}")
        else:
            print(f"\n✗ Falha na tentativa {tentativa}")
        
        if tentativa < (limite_tentativas or float('inf')):
            print(f"\n⏳ Próxima execução em {intervalo_minutos} minuto(s)...")
            time.sleep(intervalo_minutos * 60)

# Meses disponiveis no SERPRO (portugues com acento)
MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]


def baixar_somente(ano, mes):
    """Baixa o Excel do SERPRO para 1 mes e 1 ano especificos (BRASIL INTEIRO, sem filtro de UF).

    Args:
        ano: ano (int ou str), ex: 2026
        mes: mes (int 1-12) ou nome PT ('Janeiro'..'Dezembro')

    Fluxo:
        URL: select=$::Ano%20de%20Abertura,2026&select=$::M%C3%AAs%20de%20Abertura,Janeiro
        Baixa o .xlsx e retorna o caminho do arquivo.
        Nao importa no banco - quem importa eh a funcao de producao.
    """
    print("\n" + "="*60)
    print(f"MODO: BAIXAR 1 MES")
    print("="*60)

    # Normaliza ano
    ano = str(ano).strip()
    if not ano.isdigit() or len(ano) != 4:
        print(f"✗ Ano invalido: {ano}")
        return None

    # Normaliza mes
    if isinstance(mes, int):
        if mes < 1 or mes > 12:
            print(f"✗ Mes invalido: {mes}")
            return None
        mes_nome = MESES_PT[mes - 1]
    else:
        mes_nome = str(mes).strip().capitalize()
        if mes_nome not in MESES_PT:
            # Tenta match parcial
            for m in MESES_PT:
                if m.lower().startswith(mes_nome.lower()[:3]):
                    mes_nome = m
                    break
            else:
                print(f"✗ Mes invalido: {mes}")
                return None

    print(f"Filtrando: Ano={ano} | Mes={mes_nome} | Brasil inteiro")

    url = (
        "https://dd.serpro.gov.br/publico/single/?"
        "appid=7979697b-ad3d-4b28-a5bf-9cd48ea9eae7"
        "&obj=vVDJ"
        "&opt=ctxmenu,currsel"
        f"&select=$::Ano%20de%20Abertura,{ano}"
        f"&select=$::M%C3%AAs%20de%20Abertura,{mes_nome}"
    )
    
    try:
        global driver
        driver = webdriver.Chrome(options=options)
        print("✓ Navegador aberto")
        
        print(f"\nAcessando: {url}")
        driver.get(url)
        driver.maximize_window()
        print("✓ URL acessada")
        
        salvar_snapshot("00_pagina_inicial")
        
        if not aguardar_carregamento_pagina():
            return None

        if not aguardar_elemento_qv_inner_object(intervalo=10, timeout_max=600):
            return None

        if not clicar_botao_direito_meio_tela():
            return None

        # Espera 2s para o menu de contexto abrir totalmente
        time.sleep(2)

        # Clicar em "Exportar dados" usando Selenium real (AngularJS não reage a .click() JS)
        print("\n[3] Clicando em 'Exportar dados' (click nativo)...")
        try:
            li_export_dados = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "li#export-data"))
            )
            # Pega o <span> interno com o texto (mais estável para o evento)
            span = li_export_dados.find_element(By.CSS_SELECTOR, "span.lui-list__text")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", li_export_dados)
            time.sleep(0.3)
            # Clique duplo: ActionChains dispara mousedown+mouseup+click que o AngularJS escuta
            ActionChains(driver).move_to_element(span).pause(0.1).click(span).perform()
            print("✓ Clicado em 'Exportar dados' (ActionChains)")
            time.sleep(2)  # espera diálogo abrir
            salvar_snapshot("03_exportar_dados")

            # VERIFICAÇÃO: o menu de contexto deve ter sumido e o diálogo deve ter aparecido
            contexto_ainda_visivel = driver.execute_script("""
                const menu = document.querySelector('li#export-data');
                return menu && menu.offsetParent !== null;
            """)
            if contexto_ainda_visivel:
                print("⚠ Menu de contexto AINDA visível após o clique. Salvando snapshot de debug...")
                salvar_snapshot("03_debug_menu_ainda_visivel")
                return None
            print("✓ Menu de contexto sumiu (clique confirmado)")
        except Exception as e:
            print(f"✗ Falha ao clicar em 'Exportar dados': {e}")
            salvar_snapshot("03_erro_exportar_dados")
            return None

        # Espera 2s para o diálogo de exportação abrir
        time.sleep(2)

        # Clicar em "Exportar" no diálogo - agora com click nativo também
        print("\n[4] Clicando em 'Exportar' (botão do diálogo)...")
        try:
            # Tenta achar pelo texto exato em <button> do Qlik
            botao_exportar = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='Exportar' or .//span[normalize-space(.)='Exportar']]"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", botao_exportar)
            time.sleep(0.3)
            ActionChains(driver).move_to_element(botao_exportar).pause(0.1).click(botao_exportar).perform()
            print(f"✓ Clicado em 'Exportar' (botão nativo)")
            time.sleep(2)
            salvar_snapshot("04_exportar")

            # VERIFICAÇÃO: o diálogo deve ter sumido e o link de download deve começar a aparecer
            # (link pode demorar até 180s, então só checamos que o diálogo sumiu)
            dialogo_ainda_visivel = driver.execute_script("""
                const btn = Array.from(document.querySelectorAll('button, span'))
                    .find(el => (el.innerText || el.textContent || '').trim() === 'Exportar' && el.offsetParent);
                return btn !== undefined;
            """)
            if dialogo_ainda_visivel:
                print("⚠ Botão 'Exportar' AINDA visível após o clique. Salvando snapshot de debug...")
                salvar_snapshot("04_debug_botao_ainda_visivel")
                return None
            print("✓ Diálogo fechou (clique confirmado)")
        except Exception as e:
            print(f"✗ 'Exportar' (botão final) não encontrado: {e}")
            salvar_snapshot("04_erro_exportar")
            return None

        # Espera 2s antes de procurar o link de download
        time.sleep(2)

        # Aguardar link de download aparecer
        print("\n[5] Aguardando link de download...")
        if not aguardar_botao_export_url(intervalo=5, timeout_max=180):
            return None

        # Aguardar download terminar
        print("\nAguardando finalizacao do download...")
        if not aguardar_download_terminar(timeout_max=600):
            print("\n✗ Download nao completou")
            return None
        
        arquivos = [f for f in os.listdir(download_folder) if f.endswith('.xlsx') and not f.endswith('.crdownload')]
        if arquivos:
            # Pega o mais recente por mtime
            arquivos.sort(key=lambda f: os.path.getmtime(os.path.join(download_folder, f)), reverse=True)
            arquivo_baixado = os.path.join(download_folder, arquivos[0])
            print(f"\n✓ Download concluido: {arquivo_baixado}")
            return arquivo_baixado
        else:
            print("\n✗ Nenhum arquivo baixado encontrado")
            return None

    except Exception as e:
        print(f"\n✗ Erro: {e}")
        salvar_snapshot("99_erro_geral")
        return None
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
            print("✓ Navegador fechado")


def aguardar_download_terminar(intervalo=5, timeout_max=600):
    """Aguarda o arquivo .crdownload sumir (download completo)"""
    print(f"Aguardando download (timeout: {timeout_max}s)...")
    tempo_decorrido = 0
    
    while tempo_decorrido < timeout_max:
        crdownload_files = [f for f in os.listdir(download_folder) if f.endswith('.crdownload')]
        
        if not crdownload_files:
            print(f"✓ Download completo! ({tempo_decorrido}s)")
            return True
        
        tamanho = os.path.getsize(os.path.join(download_folder, crdownload_files[0])) if crdownload_files else 0
        print(f"  Download em andamento: {crdownload_files[0]} ({tamanho/1024/1024:.1f}MB)")
        
        time.sleep(intervalo)
        tempo_decorrido += intervalo
    
    print(f"✗ Timeout no download ({timeout_max}s)")
    return False


def importar_ultimo_download(apagar_apos=True):
    """Importa o último arquivo baixado para o banco SQLite (o mais recente por mtime).

    Args:
        apagar_apos: se True (default), apaga o xlsx apos importar com sucesso
    """
    arquivos = [f for f in os.listdir(download_folder)
                if f.endswith('.xlsx') and not f.endswith('.crdownload')]

    if not arquivos:
        print("✗ Nenhum arquivo .xlsx encontrado na pasta de downloads")
        return False

    # Pega o MAIS RECENTE por mtime (evita pegar Excel antigo de runs anteriores)
    arquivos.sort(key=lambda f: os.path.getmtime(os.path.join(download_folder, f)), reverse=True)
    arquivo_baixado = os.path.join(download_folder, arquivos[0])
    print(f"\nArquivo encontrado: {arquivo_baixado}")

    # Detecta o ano atual no xlsx para reimportar (so esse ano)
    from datetime import datetime as _dt
    anos_xlsx = _detectar_anos_xlsx(arquivo_baixado)
    ano_atual = _dt.now().year
    anos_para_reimportar = [a for a in anos_xlsx if a == ano_atual] or None

    ok = importar_excel_para_sqlite(arquivo_baixado, anos_para_reimportar=anos_para_reimportar)

    if ok and apagar_apos:
        try:
            os.remove(arquivo_baixado)
            print(f"  🗑 Arquivo removido: {os.path.basename(arquivo_baixado)}")
        except Exception as e:
            print(f"  ⚠ Nao foi possivel remover: {e}")
    elif not ok:
        print(f"  ⚠ xlsx MANTIDO para inspecao (importacao falhou)")

    return ok


def executar_ciclo_completo():
    """Executa ciclo completo: baixar + importar"""
    print("\n" + "="*60)
    print("INICIANDO CICLO COMPLETO (BAIXAR + IMPORTAR)")
    print("="*60)
    
    # Etapa 1: Baixar
    if not baixar_somente():
        print("\n✗ Falha na etapa de download")
        return False
    
    # Etapa 2: Importar
    if not importar_ultimo_download():
        print("\n✗ Falha na etapa de importacao")
        return False
    
    print("\n" + "="*60)
    print("✓ CICLO COMPLETO FINALIZADO COM SUCESSO!")
    print("="*60)
    return True


def _detectar_anos_xlsx(caminho_arquivo):
    """Detecta os anos presentes em um xlsx (coluna ano_abertura).
    Le no maximo 5000 linhas para performance. Retorna set() se nao conseguir."""
    anos = set()
    try:
        wb = load_workbook(filename=caminho_arquivo, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5001, values_only=True)):
            if row and row[0] is not None:
                try:
                    anos.add(int(str(row[0]).strip()))
                except (ValueError, TypeError):
                    pass
        wb.close()
    except Exception as e:
        print(f"  ⚠ Erro ao detectar anos do xlsx: {e}")
    return anos


def arquivo_respeita_filtro(caminho_arquivo, ano_esperado, mes_esperado, max_linhas_amostra=200):
    """Verifica se o Excel baixado realmente se refere ao (ano, mes) solicitado.

    O SERPRO as vezes IGNORA o filtro de mes quando o valor nao bate
    (ex: pediu 'Dezembro/2026', mas ainda nao tem -> retorna TODOS os anos/meses).
    Esta funcao detecta isso: le as primeiras max_linhas_amostra linhas e
    checa se todas tem ano == ano_esperado e mes == mes_esperado.

    Retorna:
        True  = filtro respeitado OU arquivo vazio (sem dados)
        False = filtro ignorado (vai descartar o xlsx)
    """
    if not os.path.exists(caminho_arquivo):
        return True

    MESES_MAP = {
        "janeiro": 1, "fevereiro": 2, "marco": 3, "março": 3, "abril": 4,
        "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
        "outubro": 10, "novembro": 11, "dezembro": 12
    }

    def normaliza_mes(valor):
        """Converte valor de mes (1-12, 'Janeiro', etc) para int 1-12."""
        if valor is None:
            return None
        s = str(valor).strip().lower()
        if not s:
            return None
        # Tenta como int direto (ex: "1", "12")
        try:
            n = int(s)
            if 1 <= n <= 12:
                return n
        except (ValueError, TypeError):
            pass
        # Tenta como nome do mes
        if s in MESES_MAP:
            return MESES_MAP[s]
        # Tenta match parcial (ex: "Marco" -> "marco")
        for nome, num in MESES_MAP.items():
            if nome in s or s in nome:
                return num
        return None

    try:
        wb = load_workbook(filename=caminho_arquivo, read_only=True, data_only=True)
        ws = wb.active

        # Le cabecalho para descobrir indices das colunas
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return True  # sem header = vazio

        # Encontra indices: coluna que tem "ano" + "abertura" e "mes" + "abertura"
        idx_ano = None
        idx_mes = None
        for i, h in enumerate(header_row):
            if h is None:
                continue
            h_norm = str(h).strip().lower()
            # Remove acentos para comparacao
            h_norm = (h_norm.replace("ã", "a").replace("á", "a")
                          .replace("é", "e").replace("ê", "e").replace("í", "i")
                          .replace("ó", "o").replace("ô", "o").replace("ú", "u")
                          .replace("ç", "c"))
            if "ano" in h_norm and "abert" in h_norm and idx_ano is None:
                idx_ano = i
            elif "mes" in h_norm and "abert" in h_norm and idx_mes is None:
                idx_mes = i

        if idx_ano is None or idx_mes is None:
            wb.close()
            # Nao conseguiu mapear colunas - aceita (mas loga)
            return True

        # Conta quantas linhas tem o filtro correto
        total = 0
        corretas = 0
        anos_vistos = set()
        meses_vistos = set()
        for row in ws.iter_rows(min_row=2, max_row=max_linhas_amostra + 1, values_only=True):
            if not row or idx_ano >= len(row) or row[idx_ano] is None:
                continue
            total += 1
            try:
                a = int(str(row[idx_ano]).strip())
            except (ValueError, TypeError):
                continue
            m = None
            if idx_mes < len(row):
                m = normaliza_mes(row[idx_mes])
            anos_vistos.add(a)
            if m is not None:
                meses_vistos.add(m)
            if a == ano_esperado and m == mes_esperado:
                corretas += 1
        wb.close()

        if total == 0:
            return True  # vazio

        # Se >= 95% das linhas tem o filtro correto, aceita
        aceita = (corretas / total) >= 0.95
        if not aceita:
            print(f"  📊 Amostra: {corretas}/{total} corretas")
            print(f"     Anos vistos: {sorted(anos_vistos)[:10]}{'...' if len(anos_vistos)>10 else ''}")
            print(f"     Meses vistos: {sorted(meses_vistos)}")
            print(f"     Esperado: ano={ano_esperado}, mes={mes_esperado}")
        return aceita
    except Exception as e:
        print(f"  ⚠ Erro ao verificar filtro: {e}")
        return True  # em caso de erro, aceita (nao bloqueia)


def arquivo_eh_vazio(caminho_arquivo, max_bytes=50000):
    """Detecta se o Excel baixado está vazio/sem dados.

    O SERPRO retorna um Excel com apenas cabeçalho (ou quase vazio) quando
    a combinação UF+anos não tem registros. Critérios:
      1) arquivo não existe ou tem < max_bytes (geralmente < 50KB = sem dados)
      2) ou tem 0-1 linhas de dados (só cabeçalho)
    """
    if not os.path.exists(caminho_arquivo):
        return True
    tamanho = os.path.getsize(caminho_arquivo)
    if tamanho < max_bytes:
        return True
    try:
        wb = load_workbook(filename=caminho_arquivo, read_only=True, data_only=True)
        ws = wb.active
        # max_row conta linhas usadas; se for <= 1 só tem cabeçalho
        n_linhas = ws.max_row or 0
        wb.close()
        if n_linhas <= 1:
            return True
    except Exception as e:
        print(f"  ⚠ Erro ao inspecionar Excel: {e}")
        return False
    return False


def ja_tem_no_banco(ano, mes, minimo_registros=10):
    """Verifica se o (ano, mes) ja foi importado no banco SQLite.

    Robusto contra variacoes:
        - mes pode ser int (3) ou string ('3', 'Marco', 'Marco')
        - o banco pode ter encoding baguncado ('Mar�o' em vez de 'Marco')
        - acento/mes escritos de formas diferentes

    Retorna:
        (True, N)  se ja existem >= minimo_registros registros deste (ano, mes)
        (False, 0) caso contrario (ou erro)
    """
    MESES_NOMES = {
        1: ["janeiro", "jan"], 2: ["fevereiro", "fev"],
        3: ["marco", "mar", "março"], 4: ["abril", "abr"],
        5: ["maio", "mai"], 6: ["junho", "jun"],
        7: ["julho", "jul"], 8: ["agosto", "ago"],
        9: ["setembro", "set"], 10: ["outubro", "out"],
        11: ["novembro", "nov"], 12: ["dezembro", "dez"]
    }

    def normaliza(s):
        """Remove acento e baixa caixa."""
        if s is None:
            return ""
        s = str(s).strip().lower()
        # Remove acentos
        s = (s.replace("ã", "a").replace("á", "a").replace("à", "a")
              .replace("é", "e").replace("ê", "e")
              .replace("í", "i")
              .replace("ó", "o").replace("ô", "o").replace("õ", "o")
              .replace("ú", "u").replace("ç", "c"))
        return s

    # Descobre o numero do mes
    if isinstance(mes, int):
        mes_num = mes
    else:
        s = normaliza(mes)
        if s.isdigit():
            mes_num = int(s)
        else:
            mes_num = None
            for n, nomes in MESES_NOMES.items():
                for nome in nomes:
                    if nome in s or s in nome:
                        mes_num = n
                        break
                if mes_num:
                    break
            if mes_num is None:
                # Sem mes reconhecido, considera que NAO tem (melhor tentar baixar)
                return (False, 0)

    nomes_mes = MESES_NOMES.get(mes_num, [])
    # Lista de padroes aceitos para match
    padroes = [str(mes_num), f"{mes_num:02d}"] + nomes_mes

    try:
        cnxn = sqlite3.connect('base_dados.db')
        cur = cnxn.cursor()
        # Tenta match exato primeiro (caso feliz)
        for padrao in padroes:
            cur.execute(
                "SELECT COUNT(*) FROM dados_serpro WHERE ano_abertura = ? AND mes_abertura = ?",
                (str(ano), padrao)
            )
            n = cur.fetchone()[0]
            if n > 0:
                cnxn.close()
                return (n >= minimo_registros, n)
        # Fallback: normaliza tudo no Python e compara
        cur.execute(
            "SELECT COUNT(*) FROM dados_serpro WHERE ano_abertura = ?",
            (str(ano),)
        )
        # Pega os meses distintos deste ano e compara normalizado
        cur.execute(
            "SELECT DISTINCT mes_abertura FROM dados_serpro WHERE ano_abertura = ?",
            (str(ano),)
        )
        for (m_banco,) in cur.fetchall():
            if normaliza(m_banco) in [normaliza(p) for p in padroes]:
                cur.execute(
                    "SELECT COUNT(*) FROM dados_serpro WHERE ano_abertura = ? AND mes_abertura = ?",
                    (str(ano), m_banco)
                )
                n = cur.fetchone()[0]
                cnxn.close()
                return (n >= minimo_registros, n)
        cnxn.close()
        return (False, 0)
    except Exception as e:
        print(f"  ⚠ Erro ao consultar banco: {e}")
        return (False, 0)


def producao_completa_brasil(ano_inicio=None, ano_fim=None,
                              intervalo_minutos=0, limite_meses=None,
                              mes_inicio=None, mes_fim=None):
    """Producao completa do Brasil inteiro, 1 mes e 1 ano por vez.

    Varre de (ano_fim, mes_inicio) ate (ano_inicio, mes_fim) em ordem
    DECRESCENTE. Para cada combinacao:
        1. Baixa o Excel do SERPRO (Brasil inteiro, filtro Ano+Mes)
        2. Detecta se veio vazio
        3. Detecta se o SERPRO IGNOROU o filtro (caso comum: pediu mes
           que ainda nao existe e o SERPRO retorna TUDO) - nesse caso
           descarta o xlsx e segue sem contar
        4. Importa com INSERT OR IGNORE (acumula, nao duplica)
        5. Apaga o .xlsx

    Criterio de parada: chegar em 1932 (ano_inicio) OU 3 anos consecutivos
    sem dados (ano inteiro vazio, o que indica fim real do historico).

    Args:
        ano_inicio: ano mais antigo (default 1932 - SERPRO tem dados antigos)
        ano_fim: ano mais recente (default ano atual)
        intervalo_minutos: espera entre cada download (default 0)
        limite_meses: para apos N meses processados (None = sem limite)
        mes_inicio: mes de inicio 1-12 (default 12 = mais recente primeiro)
        mes_fim: mes final 1-12 (default 1 = janeiro)
    """
    from datetime import datetime
    agora = datetime.now()
    mes_atual = agora.month
    if ano_fim is None:
        ano_fim = agora.year
    if ano_inicio is None:
        ano_inicio = 1932  # SERPRO tem dados antigos
    # IMPORTANTE: o mes maximo que pode ter dados no SERPRO e o mes
    # ANTERIOR ao atual. Em Junho, o maximo e Maio (mes_atual - 1).
    # Em Janeiro, o maximo e Dezembro do ano anterior (ano_fim - 1, mes 12).
    if mes_inicio is None:
        mes_inicio = mes_atual - 1
    else:
        # Respeita o argumento, mas se for maior que o mes maximo permitido, limita
        if mes_inicio > mes_atual - 1:
            mes_inicio = mes_atual - 1
    if mes_fim is None:
        mes_fim = 1
    # Se mes_inicio ficar 0 (estamos em Janeiro), o ultimo mes valido e
    # Dezembro do ano ANTERIOR.
    if mes_inicio < 1:
        mes_inicio = 12
        ano_fim = agora.year - 1

    print("\n" + "="*70)
    print("PRODUCAO COMPLETA BRASIL - 1 MES POR VEZ")
    print(f"  Hoje: {MESES_PT[mes_atual-1]}/{agora.year}")
    print(f"  Mes maximo permitido: {MESES_PT[mes_inicio-1]}/{ano_fim} (mes atual - 1)")
    print(f"  Range: {ano_inicio}..{ano_fim} | Meses: {mes_inicio}..{mes_fim}")
    print(f"  Intervalo: {intervalo_minutos} min | Limite: {limite_meses or 'infinito'} meses")
    print(f"  Brasil inteiro (sem filtro de UF)")
    print(f"  Ordem: do mais RECENTE para o mais ANTIGO")
    print("="*70)

    # Gera lista de (ano, mes) em ordem DECRESCENTE
    # - Para o ano_fim: meses de mes_inicio ate 1 (decrescente, limitado a mes_atual)
    # - Para anos intermediarios: todos os 12 meses (decrescente)
    # - Para o ano_inicio: meses de 12 ate mes_fim (decrescente)
    combinacoes = []
    for ano in range(ano_fim, ano_inicio - 1, -1):
        if ano == ano_fim:
            meses_ano = range(mes_inicio, 0, -1)
        elif ano == ano_inicio:
            meses_ano = range(12, mes_fim - 1, -1)
        else:
            meses_ano = range(12, 0, -1)
        for mes in meses_ano:
            combinacoes.append((ano, mes))

    total_baixados = 0
    total_importados = 0
    total_vazios = 0
    total_filtro_ignorado = 0
    anos_vazios_consecutivos = 0
    ultimo_ano_com_dados = None
    mes_num = 0
    inicio = time.time()

    for ano, mes in combinacoes:
        mes_num += 1

        if limite_meses and mes_num > limite_meses:
            print(f"\n✓ Limite de {limite_meses} meses atingido. Encerrando.")
            break

        elapsed = time.time() - inicio
        print(f"\n>>> [{mes_num}/{len(combinacoes)}] {MESES_PT[mes-1]}/{ano} (decorrido: {elapsed/60:.1f} min)")
        print("-" * 70)

        # 0) VALIDACAO: ja tem no banco? Se sim, pula
        tem, n_registros = ja_tem_no_banco(ano, mes)
        if tem:
            print(f"  ⏭  Ja existe no banco ({n_registros:,} registros). Pulando.".replace(',', '.'))
            continue

        # 1) Baixar
        caminho = baixar_somente(ano=ano, mes=mes)
        if not caminho:
            print(f"  ✗ Falha no download de {MESES_PT[mes-1]}/{ano}. Continuando...")
            continue

        # 2) Verificar se o filtro foi respeitado
        # Se o SERPRO retornou dados de outros anos/meses, descarta
        if not arquivo_respeita_filtro(caminho, ano, mes):
            print(f"  ⚠ Filtro IGNORADO pelo SERPRO (pediu {MESES_PT[mes-1]}/{ano} mas retornou dados de outros periodos).")
            total_filtro_ignorado += 1
            try:
                os.remove(caminho)
                print(f"  🗑 Removido: {os.path.basename(caminho)}")
            except Exception:
                pass
            continue  # nao conta como vazio do historico

        # 3) Verificar se veio vazio
        if arquivo_eh_vazio(caminho):
            print(f"  ⏭ Arquivo VAZIO (sem dados para {MESES_PT[mes-1]}/{ano}).")
            total_vazios += 1
            try:
                os.remove(caminho)
                print(f"  🗑 Removido: {os.path.basename(caminho)}")
            except Exception:
                pass
            # 3 meses vazios consecutivos = provavelmente fim do historico deste periodo
            # Mas so paramos se 3 ANOS inteiros vazios (o que so vai acontecer qnd chegarmos
            # num periodo que o SERPRO nao tem). Continua mesmo assim.
            continue

        # 4) Importar
        total_vazios = 0
        total_baixados += 1
        ultimo_ano_com_dados = ano

        if importar_excel_para_sqlite(caminho, limpar_antes=False, anos_para_reimportar=None):
            total_importados += 1
            print(f"  ✓ {MESES_PT[mes-1]}/{ano} OK")
            # Apaga o xlsx
            try:
                os.remove(caminho)
                print(f"  🗑 Arquivo removido: {os.path.basename(caminho)}")
            except Exception as e:
                print(f"  ⚠ Nao foi possivel remover: {e}")
        else:
            print(f"  ✗ Falha na importacao de {MESES_PT[mes-1]}/{ano} (xlsx MANTIDO)")

        # 5) Intervalo entre downloads
        if intervalo_minutos > 0 and mes_num < len(combinacoes):
            if limite_meses is None or mes_num < limite_meses:
                print(f"\n⏳ Aguardando {intervalo_minutos} min antes do proximo mes...")
                time.sleep(intervalo_minutos * 60)

        # 4) Intervalo entre downloads (se houver mais)
        if intervalo_minutos > 0 and mes_num < len(combinacoes):
            if limite_meses is None or mes_num < limite_meses:
                print(f"\n⏳ Aguardando {intervalo_minutos} min antes do proximo mes...")
                time.sleep(intervalo_minutos * 60)

    # Relatorio final
    elapsed = time.time() - inicio
    print("\n" + "="*70)
    print("RESUMO DA PRODUCAO")
    print("="*70)
    print(f"  Meses processados:   {mes_num}")
    print(f"  Downloads OK:        {total_baixados}")
    print(f"  Importacoes OK:      {total_importados}")
    print(f"  Meses vazios:        {total_vazios}")
    print(f"  Tempo total:         {elapsed/60:.1f} min")
    # Tamanho do banco final
    try:
        cnxn = sqlite3.connect('base_dados.db')
        cur = cnxn.cursor()
        cur.execute("SELECT COUNT(*) FROM dados_serpro")
        total_banco = cur.fetchone()[0]
        cur.execute("SELECT MIN(ano_abertura), MAX(ano_abertura) FROM dados_serpro")
        anos_banco = cur.fetchone()
        cnxn.close()
        print(f"  Banco: {total_banco:,} registros | Anos: {anos_banco[0]}..{anos_banco[1]}".replace(',', '.'))
    except Exception as e:
        print(f"  (erro ao consultar banco: {e})")
    print("="*70)
    return total_importados > 0


# Aliases para compatibilidade
importar_completo_brasil = producao_completa_brasil


if __name__ == "__main__":
    import argparse
    
    print("EXPORTADOR DE DADOS SERPRO COM SNAPSHOTS")
    print("========================================\n")
    
    parser = argparse.ArgumentParser(description="Exportador SERPRO - Mapa de Empresas")
    parser.add_argument("acao", nargs="?", default="producao-completa",
                        choices=["producao-completa", "producao-completa-brasil",
                                 "importar-completo-brasil", "baixar-mes", "importar", "info"],
                        help=("Acoes disponiveis: "
                              "producao-completa (padrao, baixa+importa BRASIL mes a mes), "
                              "baixar-mes (baixa 1 mes especifico), "
                              "importar (importa o xlsx mais recente), "
                              "info (status do banco)"))
    parser.add_argument("--ano", "-y", type=int, default=None,
                        help="(baixar-mes) Ano (ex: 2026). Default: ano atual")
    parser.add_argument("--mes", "-m", default=None,
                        help="(baixar-mes) Mes: 1-12 ou nome PT (Janeiro, Fevereiro, ...). Default: mes atual")
    parser.add_argument("--arquivo", "-f", type=str, default=None,
                        help="(importar) Caminho do arquivo .xlsx")
    parser.add_argument("--ano-inicio", type=int, default=None,
                        help="(producao-completa) Ano inicial (default: 1932)")
    parser.add_argument("--ano-fim", type=int, default=None,
                        help="(producao-completa) Ano final (default: ano atual)")
    parser.add_argument("--mes-inicio", type=int, default=None,
                        help="(producao-completa) Mes inicial - ordem decrescente (default: mes atual - 1, ex: Junho -> Maio)")
    parser.add_argument("--mes-fim", type=int, default=1,
                        help="(producao-completa) Mes final - ordem decrescente (default: 1)")
    parser.add_argument("--intervalo", type=int, default=0,
                        help="(producao-completa) Minutos de espera entre meses (default 0)")
    parser.add_argument("--limite-meses", type=int, default=None,
                        help="(producao-completa) Para apos N meses (default: ate 3 vazios consecutivos)")
    parser.add_argument("--exemplos", action="store_true",
                        help="Mostra exemplos de uso e sai")

    args = parser.parse_args()

    if args.exemplos:
        print("""
============================================================
EXEMPLOS DE USO - Exportador SERPRO (Mapa de Empresas)
============================================================

URL base montada (Brasil inteiro, 1 mes especifico):
  https://dd.serpro.gov.br/publico/single/?appid=...&obj=vVDJ
  &opt=ctxmenu,currsel
  &select=$::Ano%20de%20Abertura,2026
  &select=$::M%C3%AAs%20de%20Abertura,Janeiro

1) PRODUCAO COMPLETA BRASIL: baixa+importa+apaga 1 mes por vez, do mais recente:
   python exportador_dados_serpro.py
   (mesmo que: python exportador_dados_serpro.py producao-completa)

2) Producao completa com range customizado (ex: 2015 a 2026):
   python exportador_dados_serpro.py --ano-inicio 2015 --ano-fim 2026

3) Producao completa com 2 min entre cada mes (sem sobrecarregar):
   python exportador_dados_serpro.py --intervalo 2

4) Testar com apenas 2 meses (rapido):
   python exportador_dados_serpro.py --limite-meses 2

5) Baixar apenas 1 mes especifico (teste):
   python exportador_dados_serpro.py baixar-mes --ano 2026 --mes 1
   python exportador_dados_serpro.py baixar-mes -y 2025 -m Marco

6) Importar um .xlsx especifico:
   python exportador_dados_serpro.py importar -f "downloads_serpro/foo.xlsx"

7) Status do banco e downloads:
   python exportador_dados_serpro.py info
============================================================
""")
        sys.exit(0)
    
    if args.acao == "info":
        print("=== STATUS DO SISTEMA ===")
        arquivos = [f for f in os.listdir(download_folder) if f.endswith('.xlsx') and not f.endswith('.crdownload')]
        if arquivos:
            # Ordena por mtime desc (mais recente primeiro)
            arquivos.sort(key=lambda f: os.path.getmtime(os.path.join(download_folder, f)), reverse=True)
            print(f"Arquivos disponiveis para importacao (mais recente primeiro):")
            for f in arquivos:
                tamanho = os.path.getsize(os.path.join(download_folder, f))
                mtime = datetime.fromtimestamp(os.path.getmtime(os.path.join(download_folder, f)))
                print(f"  - {f} ({tamanho/1024/1024:.2f} MB) - {mtime:%Y-%m-%d %H:%M:%S}")
        else:
            print("Nenhum arquivo .xlsx na pasta de downloads")
        
        try:
            cnxn = sqlite3.connect('base_dados.db')
            cursor = cnxn.cursor()
            cursor.execute("SELECT COUNT(*) FROM dados_serpro")
            total = cursor.fetchone()[0]
            cursor.execute("SELECT MIN(ano_abertura), MAX(ano_abertura) FROM dados_serpro")
            range_anos = cursor.fetchone()
            print(f"\nBanco de dados:")
            print(f"  Total registros: {total}")
            print(f"  Range anos: {range_anos[0]} a {range_anos[1]}")
            cursor.close()
            cnxn.close()
        except Exception as e:
            print(f"  Erro ao verificar banco: {e}")
    
    elif args.acao in ("producao-completa", "producao-completa-brasil", "importar-completo-brasil"):
        # Producao: BRASIL inteiro, 1 mes por vez, do mais recente para o mais antigo
        producao_completa_brasil(
            ano_inicio=args.ano_inicio,
            ano_fim=args.ano_fim,
            intervalo_minutos=args.intervalo,
            limite_meses=args.limite_meses,
            mes_inicio=args.mes_inicio,
            mes_fim=args.mes_fim,
        )

    elif args.acao == "baixar-mes":
        # Baixa 1 mes especifico
        from datetime import datetime as _dt
        agora = _dt.now()
        ano = args.ano if args.ano is not None else agora.year
        if args.mes is not None:
            try:
                mes = int(args.mes)
            except ValueError:
                mes = args.mes  # nome do mes
        else:
            mes = agora.month

        # Validacao: ja tem no banco?
        tem, n_registros = ja_tem_no_banco(ano, mes)
        if tem:
            print(f"\n⏭  Ja existe no banco: {n_registros:,} registros para {ano}/{mes}".replace(',', '.'))
            print(f"   Pulando download (force=True se quiser baixar de novo).")
            import sys
            sys.exit(0)

        caminho = baixar_somente(ano=ano, mes=mes)
        if caminho:
            print(f"\nArquivo baixado: {caminho}")
            # Importa direto
            importar_excel_para_sqlite(caminho, limpar_antes=False)
            try:
                os.remove(caminho)
                print(f"  🗑 Arquivo removido: {os.path.basename(caminho)}")
            except Exception as e:
                print(f"  ⚠ Nao foi possivel remover: {e}")

    elif args.acao == "importar":
        if args.arquivo:
            importar_excel_para_sqlite(args.arquivo)
        else:
            importar_ultimo_download()

    else:
        parser.print_help()
